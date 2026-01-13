"""
Excel export utilities.
Xuất báo cáo Excel với sheet riêng cho mỗi product và sheet Portfolio tổng hợp.
"""
import pandas as pd
import numpy as np
from typing import Optional, List
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


def _format_mixed_sheet(worksheet, segment_key: str, del_type: str, data_df: pd.DataFrame, flags_df: pd.DataFrame):
    """
    Format Mixed sheet với các yêu cầu:
    - Title tại dòng 1: {segment_key}_{del_type} Actual & Forecast
    - Header từ dòng 3, tô đậm, tô cell, căn giữa
    - Values format 4 decimal places với %
    - Color scale từ MOB 0 đến MOB cuối
    - Border phân biệt actual/forecast
    - Bỏ grid
    """
    # 1. Title tại dòng 1
    title = f"{segment_key}_{del_type} Actual & Forecast"
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=20, bold=True, color="1F4E79")  # Dark Blue
    
    # Merge cells cho title (A1 đến cột cuối)
    max_col = len(data_df.columns)
    worksheet.merge_cells(f'A1:{get_column_letter(max_col)}1')
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # 2. Headers từ dòng 3
    header_row = 3
    for col_idx, col_name in enumerate(data_df.columns, 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
        cell.alignment = Alignment(horizontal='center')
    
    # 3. Data từ dòng 4
    data_start_row = 4
    mob_cols = [col for col in data_df.columns if col.startswith('MOB_')]
    
    for row_idx, (_, row_data) in enumerate(data_df.iterrows(), data_start_row):
        for col_idx, col_name in enumerate(data_df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = row_data[col_name]
            
            if col_name in mob_cols and pd.notna(value):
                # Format as percentage with 4 decimal places
                cell.value = float(value)
                cell.number_format = '0.0000%'
            else:
                cell.value = value
    
    # 4. Color scale cho MOB columns
    if mob_cols:
        mob_start_col = data_df.columns.get_loc(mob_cols[0]) + 1
        mob_end_col = data_df.columns.get_loc(mob_cols[-1]) + 1
        
        color_scale_range = f"{get_column_letter(mob_start_col)}{data_start_row}:{get_column_letter(mob_end_col)}{data_start_row + len(data_df) - 1}"
        
        # Red-Yellow-Green color scale
        color_scale = ColorScaleRule(
            start_type='min', start_color='F8696B',  # Red
            mid_type='percentile', mid_value=50, mid_color='FFEB9C',  # Yellow
            end_type='max', end_color='63BE7B'  # Green
        )
        worksheet.conditional_formatting.add(color_scale_range, color_scale)
    
    # 5. Borders
    # Border bình thường cho tất cả cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells first
    for row_idx in range(header_row, data_start_row + len(data_df)):
        for col_idx in range(1, len(data_df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # Add thick red borders to separate actual/forecast (if flags available)
    if flags_df is not None and mob_cols:
        thick_red_side = Side(style='thick', color='FF0000')
        
        for row_idx in range(data_start_row, data_start_row + len(data_df)):
            data_row_idx = row_idx - data_start_row
            if data_row_idx >= len(flags_df):
                continue
                
            for col_idx, col_name in enumerate(data_df.columns, 1):
                if col_name not in mob_cols:
                    continue
                    
                cell = worksheet.cell(row=row_idx, column=col_idx)
                flag_value = flags_df.iloc[data_row_idx][col_name] if col_name in flags_df.columns else None
                
                if flag_value == 'FORECAST':
                    # Add thick red bottom and left borders for forecast cells
                    current_border = cell.border
                    cell.border = Border(
                        left=thick_red_side,
                        right=current_border.right,
                        top=current_border.top,
                        bottom=thick_red_side
                    )
    
    # 6. Bỏ grid
    worksheet.sheet_view.showGridLines = False
    
    # 7. Auto-fit columns
    for col_idx in range(1, len(data_df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].auto_size = True
    """Sanitize sheet name for Excel (max 31 chars, no special chars)."""
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name[:max_len]


def _sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    """Sanitize sheet name for Excel (max 31 chars, no special chars)."""
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name[:max_len]


def _ensure_unique_names(names: list) -> list:
    """Ensure sheet names are unique by appending numbers if needed."""
    seen = {}
    result = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name[:28]}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


def _format_standard_sheet(worksheet, data_df: pd.DataFrame, sheet_title: str = None):
    """
    Format standard sheet (non-Mixed sheets) với định dạng cơ bản:
    - Values format 4 decimal places với %
    - Headers tô đậm
    - Bỏ grid
    """
    # Title nếu có
    start_row = 1
    if sheet_title:
        worksheet['A1'] = sheet_title
        worksheet['A1'].font = Font(size=16, bold=True, color="1F4E79")
        start_row = 3
    
    # Headers
    header_row = start_row
    for col_idx, col_name in enumerate(data_df.columns, 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    data_start_row = header_row + 1
    mob_cols = [col for col in data_df.columns if col.startswith('MOB_')]
    
    for row_idx, (_, row_data) in enumerate(data_df.iterrows(), data_start_row):
        for col_idx, col_name in enumerate(data_df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = row_data[col_name]
            
            if col_name in mob_cols and pd.notna(value):
                cell.value = float(value)
                cell.number_format = '0.0000%'
            else:
                cell.value = value
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx in range(header_row, data_start_row + len(data_df)):
        for col_idx in range(1, len(data_df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # Bỏ grid
    worksheet.sheet_view.showGridLines = False


def _compute_portfolio_del(
    mixed_wide: pd.DataFrame,
    actual_wide: pd.DataFrame,
    forecast_wide: pd.DataFrame,
    flags_wide: pd.DataFrame
) -> tuple:
    """
    Tính DEL cho Portfolio (tổng hợp tất cả products).
    Compute Portfolio DEL by aggregating all products.
    
    Returns:
        Tuple of (portfolio_mixed, portfolio_actual, portfolio_forecast, portfolio_flags)
    """
    mob_cols = [c for c in mixed_wide.columns if c.startswith("MOB_")]
    
    # Portfolio = mean across all segments per cohort
    # (Hoặc có thể dùng weighted average nếu có EAD weights)
    
    portfolio_mixed = mixed_wide.groupby("cohort")[mob_cols].mean().reset_index()
    portfolio_mixed.insert(1, "segment_key", "PORTFOLIO")
    
    portfolio_actual = actual_wide.groupby("cohort")[mob_cols].mean().reset_index()
    portfolio_actual.insert(1, "segment_key", "PORTFOLIO")
    
    portfolio_forecast = forecast_wide.groupby("cohort")[mob_cols].mean().reset_index()
    portfolio_forecast.insert(1, "segment_key", "PORTFOLIO")
    
    # Flags: ACTUAL nếu tất cả segments có ACTUAL, else FORECAST
    def agg_flags(group):
        result = {}
        for col in mob_cols:
            vals = group[col].values
            if all(v == "ACTUAL" for v in vals if v):
                result[col] = "ACTUAL"
            elif any(v == "ACTUAL" for v in vals if v):
                result[col] = "MIXED"
            else:
                result[col] = "FORECAST"
        return pd.Series(result)
    
    portfolio_flags = flags_wide.groupby("cohort").apply(agg_flags).reset_index()
    portfolio_flags.insert(1, "segment_key", "PORTFOLIO")
    
    return portfolio_mixed, portfolio_actual, portfolio_forecast, portfolio_flags


def _split_by_segment(
    df: pd.DataFrame,
    segment_col: str = "segment_key"
) -> dict:
    """
    Chia DataFrame theo segment_key thành dict.
    Split DataFrame by segment_key into dict.
    
    Returns:
        Dict[segment_key] -> DataFrame (without segment_key column)
    """
    if segment_col not in df.columns:
        return {"ALL": df}
    
    result = {}
    for seg_key, grp in df.groupby(segment_col):
        # Drop segment_key column for cleaner output
        grp_clean = grp.drop(columns=[segment_col]).reset_index(drop=True)
        result[str(seg_key)] = grp_clean
    
    return result


def export_to_excel(
    path: str,
    transitions_long_df: pd.DataFrame,
    mixed_wide: pd.DataFrame,
    flags_wide: pd.DataFrame,
    actual_wide: pd.DataFrame,
    forecast_wide: pd.DataFrame,
    factors_df: Optional[pd.DataFrame] = None,
    forecast_df: Optional[pd.DataFrame] = None,
    meta_df: Optional[pd.DataFrame] = None
) -> None:
    """
    Export results to Excel workbook with separate sheets per product.
    Xuất báo cáo Excel với sheet riêng cho mỗi product và sheet Portfolio.
    
    Sheet structure:
    - Portfolio_Mixed, Portfolio_Actual, Portfolio_Forecast, Portfolio_Flags
    - {Product}_Mixed, {Product}_Actual, {Product}_Forecast, {Product}_Flags (per product)
    - transitions_long, segment_meta, calibration_factors, forecast_long (metadata)
    
    Args:
        path: Output file path.
        transitions_long_df: Long-form transition probabilities.
        mixed_wide: Mixed DEL report (actual where available, else forecast).
        flags_wide: Flags indicating ACTUAL or FORECAST.
        actual_wide: Actual DEL values.
        forecast_wide: Forecast DEL values.
        factors_df: Optional calibration factors.
        forecast_df: Optional long-form forecast.
        meta_df: Optional segment metadata.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        
        # === 1. PORTFOLIO SHEETS (tổng hợp tất cả products) ===
        portfolio_mixed, portfolio_actual, portfolio_forecast, portfolio_flags = \
            _compute_portfolio_del(mixed_wide, actual_wide, forecast_wide, flags_wide)
        
        # Portfolio Mixed with special formatting
        portfolio_mixed.to_excel(writer, sheet_name="Portfolio_Mixed", index=False, startrow=2)
        worksheet = writer.sheets["Portfolio_Mixed"]
        _format_mixed_sheet(worksheet, "PORTFOLIO", "DEL", portfolio_mixed, portfolio_flags)
        
        # Other Portfolio sheets
        portfolio_actual.to_excel(writer, sheet_name="Portfolio_Actual", index=False)
        worksheet = writer.sheets["Portfolio_Actual"]
        _format_standard_sheet(worksheet, portfolio_actual, "PORTFOLIO Actual")
        
        portfolio_forecast.to_excel(writer, sheet_name="Portfolio_Forecast", index=False)
        worksheet = writer.sheets["Portfolio_Forecast"]
        _format_standard_sheet(worksheet, portfolio_forecast, "PORTFOLIO Forecast")
        
        portfolio_flags.to_excel(writer, sheet_name="Portfolio_Flags", index=False)
        worksheet = writer.sheets["Portfolio_Flags"]
        _format_standard_sheet(worksheet, portfolio_flags, "PORTFOLIO Flags")
        
        # === 2. PER-PRODUCT SHEETS ===
        mixed_by_seg = _split_by_segment(mixed_wide)
        actual_by_seg = _split_by_segment(actual_wide)
        forecast_by_seg = _split_by_segment(forecast_wide)
        flags_by_seg = _split_by_segment(flags_wide)
        
        # Get all unique segments
        all_segments = sorted(set(mixed_by_seg.keys()) | set(actual_by_seg.keys()))
        
        for seg_key in all_segments:
            seg_name = _sanitize_sheet_name(seg_key, max_len=20)
            
            # Mixed with special formatting
            if seg_key in mixed_by_seg:
                sheet_name = _sanitize_sheet_name(f"{seg_name}_Mixed")
                mixed_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                worksheet = writer.sheets[sheet_name]
                flags_data = flags_by_seg.get(seg_key, None)
                _format_mixed_sheet(worksheet, seg_key, "DEL", mixed_by_seg[seg_key], flags_data)
            
            # Actual
            if seg_key in actual_by_seg:
                sheet_name = _sanitize_sheet_name(f"{seg_name}_Actual")
                actual_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                _format_standard_sheet(worksheet, actual_by_seg[seg_key], f"{seg_key} Actual")
            
            # Forecast
            if seg_key in forecast_by_seg:
                sheet_name = _sanitize_sheet_name(f"{seg_name}_Forecast")
                forecast_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                _format_standard_sheet(worksheet, forecast_by_seg[seg_key], f"{seg_key} Forecast")
            
            # Flags
            if seg_key in flags_by_seg:
                sheet_name = _sanitize_sheet_name(f"{seg_name}_Flags")
                flags_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                _format_standard_sheet(worksheet, flags_by_seg[seg_key], f"{seg_key} Flags")
        
        # === 3. METADATA SHEETS ===
        if transitions_long_df is not None:
            transitions_long_df.to_excel(writer, sheet_name="transitions_long", index=False)
            worksheet = writer.sheets["transitions_long"]
            _format_standard_sheet(worksheet, transitions_long_df, "Transition Matrices")
        
        if meta_df is not None:
            meta_df.to_excel(writer, sheet_name="segment_meta", index=False)
            worksheet = writer.sheets["segment_meta"]
            _format_standard_sheet(worksheet, meta_df, "Segment Metadata")
        
        if factors_df is not None:
            factors_df.to_excel(writer, sheet_name="calibration_factors", index=False)
            worksheet = writer.sheets["calibration_factors"]
            _format_standard_sheet(worksheet, factors_df, "Calibration Factors")
        
        if forecast_df is not None:
            forecast_df.to_excel(writer, sheet_name="forecast_long", index=False)
            worksheet = writer.sheets["forecast_long"]
            _format_standard_sheet(worksheet, forecast_df, "Forecast Data")



def export_all_del_to_excel(
    path: str,
    transitions_long_df: pd.DataFrame,
    del_results: dict,
    factors_df: Optional[pd.DataFrame] = None,
    forecast_df: Optional[pd.DataFrame] = None,
    meta_df: Optional[pd.DataFrame] = None
) -> None:
    """
    Export DEL30, DEL60, DEL90 results to Excel with separate sheets per product.
    Xuất báo cáo DEL30, DEL60, DEL90 với sheet riêng cho mỗi product và Portfolio.
    
    Sheet structure:
    - DEL30_Portfolio, DEL30_{Product}_Mixed, DEL30_{Product}_Actual, etc.
    - DEL60_Portfolio, DEL60_{Product}_Mixed, DEL60_{Product}_Actual, etc.
    - DEL90_Portfolio, DEL90_{Product}_Mixed, DEL90_{Product}_Actual, etc.
    - Metadata sheets: transitions_long, segment_meta, calibration_factors, forecast_long
    
    Args:
        path: Output file path.
        transitions_long_df: Long-form transition probabilities.
        del_results: Dict from compute_all_del_metrics with keys "DEL30", "DEL60", "DEL90"
            Each value is (mixed_wide, flags_wide, actual_wide, forecast_wide)
        factors_df: Optional calibration factors.
        forecast_df: Optional long-form forecast.
        meta_df: Optional segment metadata.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        
        # === PROCESS EACH DEL TYPE (DEL30, DEL60, DEL90) ===
        for del_type in ["del30", "del60", "del90"]:
            if del_type not in del_results:
                continue
                
            mixed_wide, flags_wide, actual_wide, forecast_wide = del_results[del_type]
            del_type_upper = del_type.upper()
            
            # === PORTFOLIO SHEETS ===
            portfolio_mixed, portfolio_actual, portfolio_forecast, portfolio_flags = \
                _compute_portfolio_del(mixed_wide, actual_wide, forecast_wide, flags_wide)
            
            # Write Portfolio Mixed sheet with formatting
            sheet_name = _sanitize_sheet_name(f"{del_type_upper}_Portfolio")
            portfolio_mixed.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            worksheet = writer.sheets[sheet_name]
            _format_mixed_sheet(worksheet, "PORTFOLIO", del_type_upper, portfolio_mixed, portfolio_flags)
            
            # Other Portfolio sheets
            sheet_name = _sanitize_sheet_name(f"{del_type_upper}_Port_Actual")
            portfolio_actual.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _format_standard_sheet(worksheet, portfolio_actual, f"PORTFOLIO {del_type_upper} Actual")
            
            sheet_name = _sanitize_sheet_name(f"{del_type_upper}_Port_Forecast")
            portfolio_forecast.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _format_standard_sheet(worksheet, portfolio_forecast, f"PORTFOLIO {del_type_upper} Forecast")
            
            sheet_name = _sanitize_sheet_name(f"{del_type_upper}_Port_Flags")
            portfolio_flags.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _format_standard_sheet(worksheet, portfolio_flags, f"PORTFOLIO {del_type_upper} Flags")
            
            # === PER-PRODUCT SHEETS ===
            mixed_by_seg = _split_by_segment(mixed_wide)
            actual_by_seg = _split_by_segment(actual_wide)
            forecast_by_seg = _split_by_segment(forecast_wide)
            flags_by_seg = _split_by_segment(flags_wide)
            
            all_segments = sorted(set(mixed_by_seg.keys()) | set(actual_by_seg.keys()))
            
            for seg_key in all_segments:
                # Truncate segment name to fit Excel limit (31 chars total)
                # Format: DEL30_TOPUP_Mixed = 16 chars max for seg_name
                seg_name = _sanitize_sheet_name(seg_key, max_len=12)
                
                # Mixed sheet with special formatting
                if seg_key in mixed_by_seg:
                    sheet_name = _sanitize_sheet_name(f"{del_type_upper}_{seg_name}_Mix")
                    mixed_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                    worksheet = writer.sheets[sheet_name]
                    flags_data = flags_by_seg.get(seg_key, None)
                    _format_mixed_sheet(worksheet, seg_key, del_type_upper, mixed_by_seg[seg_key], flags_data)
                
                # Actual
                if seg_key in actual_by_seg:
                    sheet_name = _sanitize_sheet_name(f"{del_type_upper}_{seg_name}_Act")
                    actual_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    _format_standard_sheet(worksheet, actual_by_seg[seg_key], f"{seg_key} {del_type_upper} Actual")
                
                # Forecast
                if seg_key in forecast_by_seg:
                    sheet_name = _sanitize_sheet_name(f"{del_type_upper}_{seg_name}_Fct")
                    forecast_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    _format_standard_sheet(worksheet, forecast_by_seg[seg_key], f"{seg_key} {del_type_upper} Forecast")
                
                # Flags
                if seg_key in flags_by_seg:
                    sheet_name = _sanitize_sheet_name(f"{del_type_upper}_{seg_name}_Flg")
                    flags_by_seg[seg_key].to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    _format_standard_sheet(worksheet, flags_by_seg[seg_key], f"{seg_key} {del_type_upper} Flags")
        
        # === METADATA SHEETS ===
        if transitions_long_df is not None:
            transitions_long_df.to_excel(writer, sheet_name="transitions_long", index=False)
            worksheet = writer.sheets["transitions_long"]
            _format_standard_sheet(worksheet, transitions_long_df, "Transition Matrices")
        
        if meta_df is not None:
            meta_df.to_excel(writer, sheet_name="segment_meta", index=False)
            worksheet = writer.sheets["segment_meta"]
            _format_standard_sheet(worksheet, meta_df, "Segment Metadata")
        
        if factors_df is not None:
            factors_df.to_excel(writer, sheet_name="calibration_factors", index=False)
            worksheet = writer.sheets["calibration_factors"]
            _format_standard_sheet(worksheet, factors_df, "Calibration Factors")
        
        if forecast_df is not None:
            forecast_df.to_excel(writer, sheet_name="forecast_long", index=False)
            worksheet = writer.sheets["forecast_long"]
            _format_standard_sheet(worksheet, forecast_df, "Forecast Data")
