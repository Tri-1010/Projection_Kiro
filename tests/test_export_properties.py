"""
Property-based tests for export module.

Tests the following properties from the design document:
- Property 6: Excel Percentage Format
- Property 7: ACTUAL/FORECAST Boundary Border
- Property 8: Portfolio DEL Aggregation
- Property 9: Portfolio Flag Aggregation

Validates: Requirements 4.4, 4.6, 6.1, 6.2
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
import pandas as pd
import numpy as np
from hypothesis import given, strategies as st, settings, assume
import tempfile
from openpyxl import load_workbook

from export import (
    _compute_portfolio_del,
    _format_mixed_sheet,
    export_to_excel,
)


# =============================================================================
# Test Data Generators
# =============================================================================

def create_del_dataframes(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    last_actual_mob: int,
    seed: int = None
) -> tuple:
    """
    Create synthetic DEL DataFrames for testing.
    
    Returns:
        Tuple of (mixed_wide, actual_wide, forecast_wide, flags_wide)
    """
    if seed is not None:
        np.random.seed(seed)
    
    cohorts = [f"2024-{str(i+1).zfill(2)}" for i in range(n_cohorts)]
    segments = [f"SEG_{chr(65+i)}" for i in range(n_segments)]
    mob_cols = [f"MOB_{i}" for i in range(max_mob + 1)]
    
    records_mixed = []
    records_actual = []
    records_forecast = []
    records_flags = []
    
    for cohort in cohorts:
        for segment in segments:
            mixed_row = {"cohort": cohort, "segment_key": segment}
            actual_row = {"cohort": cohort, "segment_key": segment}
            forecast_row = {"cohort": cohort, "segment_key": segment}
            flags_row = {"cohort": cohort, "segment_key": segment}
            
            for mob in range(max_mob + 1):
                mob_col = f"MOB_{mob}"
                
                # Generate random DEL values (0-30%)
                actual_val = np.random.uniform(0, 0.3) if mob <= last_actual_mob else np.nan
                forecast_val = np.random.uniform(0, 0.3)
                
                if mob <= last_actual_mob:
                    mixed_row[mob_col] = actual_val
                    flags_row[mob_col] = "ACTUAL"
                else:
                    mixed_row[mob_col] = forecast_val
                    flags_row[mob_col] = "FORECAST"
                
                actual_row[mob_col] = actual_val
                forecast_row[mob_col] = forecast_val
            
            records_mixed.append(mixed_row)
            records_actual.append(actual_row)
            records_forecast.append(forecast_row)
            records_flags.append(flags_row)
    
    mixed_wide = pd.DataFrame(records_mixed)
    actual_wide = pd.DataFrame(records_actual)
    forecast_wide = pd.DataFrame(records_forecast)
    flags_wide = pd.DataFrame(records_flags)
    
    return mixed_wide, actual_wide, forecast_wide, flags_wide


# =============================================================================
# Property 6: Excel Percentage Format
# =============================================================================
# **Feature: del-forecast-calibration, Property 6: Excel Percentage Format**
# *For any* MOB cell in the Excel output, the number format SHALL be '0.00%' (2 decimal places).
# **Validates: Requirements 4.4**

@settings(max_examples=20, deadline=60000)
@given(
    n_cohorts=st.integers(min_value=1, max_value=2),
    n_segments=st.integers(min_value=1, max_value=2),
    max_mob=st.integers(min_value=3, max_value=5),
    seed=st.integers(min_value=0, max_value=10000),
)
def test_property_6_excel_percentage_format(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    seed: int
):
    """
    Property 6: Excel Percentage Format
    
    For any MOB cell in the Excel output, the number format SHALL be '0.00%'
    (2 decimal places).
    
    **Validates: Requirements 4.4**
    """
    last_actual_mob = max_mob // 2
    
    mixed_wide, actual_wide, forecast_wide, flags_wide = create_del_dataframes(
        n_cohorts, n_segments, max_mob, last_actual_mob, seed
    )
    
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        # Export to Excel
        transitions_long_df = pd.DataFrame({"dummy": [1]})
        export_to_excel(
            tmp_path,
            transitions_long_df,
            mixed_wide,
            flags_wide,
            actual_wide,
            forecast_wide
        )
        
        # Load and verify
        wb = load_workbook(tmp_path)
        
        # Check Portfolio_Mixed sheet
        ws = wb["Portfolio_Mixed"]
        
        # Find MOB columns (starting from column 2 after cohort, segment_key)
        mob_cols = [f"MOB_{i}" for i in range(max_mob + 1)]
        
        # Data starts at row 4 (row 1=title, row 2=empty, row 3=header)
        data_start_row = 4
        
        # Check number format for MOB cells
        for row_idx in range(data_start_row, data_start_row + n_cohorts):
            for col_idx in range(3, 3 + len(mob_cols)):  # MOB columns start at column 3
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    assert cell.number_format == '0.00%', (
                        f"Cell ({row_idx}, {col_idx}) should have format '0.00%', "
                        f"got '{cell.number_format}'"
                    )
        
        wb.close()
    finally:
        # Cleanup
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# =============================================================================
# Property 7: ACTUAL/FORECAST Boundary Border
# =============================================================================
# **Feature: del-forecast-calibration, Property 7: ACTUAL/FORECAST Boundary Border**
# *For any* row in the Mixed sheet where there is a transition from ACTUAL to FORECAST,
# the last ACTUAL cell SHALL have thick red borders on its RIGHT and BOTTOM edges.
# **Validates: Requirements 4.6**

@settings(max_examples=20, deadline=60000)
@given(
    n_cohorts=st.integers(min_value=1, max_value=2),
    n_segments=st.integers(min_value=1, max_value=2),
    max_mob=st.integers(min_value=4, max_value=6),
    last_actual_mob=st.integers(min_value=1, max_value=2),
    seed=st.integers(min_value=0, max_value=10000),
)
def test_property_7_actual_forecast_boundary_border(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    last_actual_mob: int,
    seed: int
):
    """
    Property 7: ACTUAL/FORECAST Boundary Border
    
    For any row in the Mixed sheet where there is a transition from ACTUAL to FORECAST,
    the last ACTUAL cell SHALL have thick red borders on its RIGHT and BOTTOM edges.
    
    **Validates: Requirements 4.6**
    """
    assume(last_actual_mob < max_mob)  # Ensure there's a boundary
    
    mixed_wide, actual_wide, forecast_wide, flags_wide = create_del_dataframes(
        n_cohorts, n_segments, max_mob, last_actual_mob, seed
    )
    
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        # Export to Excel
        transitions_long_df = pd.DataFrame({"dummy": [1]})
        export_to_excel(
            tmp_path,
            transitions_long_df,
            mixed_wide,
            flags_wide,
            actual_wide,
            forecast_wide
        )
        
        # Load and verify
        wb = load_workbook(tmp_path)
        
        # Check Portfolio_Mixed sheet
        ws = wb["Portfolio_Mixed"]
        
        # Data starts at row 4
        data_start_row = 4
        
        # The last ACTUAL MOB column index
        # Columns: cohort(1), segment_key(2), MOB_0(3), MOB_1(4), ...
        # last_actual_mob column = 3 + last_actual_mob
        last_actual_col = 3 + last_actual_mob
        
        # Check border on last ACTUAL cell for each data row
        for row_idx in range(data_start_row, data_start_row + n_cohorts):
            cell = ws.cell(row=row_idx, column=last_actual_col)
            
            # Check right border is thick and red
            right_border = cell.border.right
            assert right_border is not None, (
                f"Cell ({row_idx}, {last_actual_col}) should have right border"
            )
            assert right_border.style == 'thick', (
                f"Cell ({row_idx}, {last_actual_col}) right border should be 'thick', "
                f"got '{right_border.style}'"
            )
            # Color can be FFFF0000 or 00FF0000 (alpha channel varies)
            assert right_border.color is not None and right_border.color.rgb.endswith('FF0000'), (
                f"Cell ({row_idx}, {last_actual_col}) right border should be red (FF0000), "
                f"got '{right_border.color.rgb}'"
            )
            
            # Check bottom border is thick and red
            bottom_border = cell.border.bottom
            assert bottom_border is not None, (
                f"Cell ({row_idx}, {last_actual_col}) should have bottom border"
            )
            assert bottom_border.style == 'thick', (
                f"Cell ({row_idx}, {last_actual_col}) bottom border should be 'thick', "
                f"got '{bottom_border.style}'"
            )
            # Color can be FFFF0000 or 00FF0000 (alpha channel varies)
            assert bottom_border.color is not None and bottom_border.color.rgb.endswith('FF0000'), (
                f"Cell ({row_idx}, {last_actual_col}) bottom border should be red (FF0000), "
                f"got '{bottom_border.color.rgb}'"
            )
        
        wb.close()
    finally:
        # Cleanup
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# =============================================================================
# Property 8: Portfolio DEL Aggregation
# =============================================================================
# **Feature: del-forecast-calibration, Property 8: Portfolio DEL Aggregation**
# *For any* MOB, the Portfolio DEL SHALL equal the mean of all segment DELs at that MOB.
# **Validates: Requirements 6.1**

@settings(max_examples=30, deadline=30000)
@given(
    n_cohorts=st.integers(min_value=1, max_value=3),
    n_segments=st.integers(min_value=2, max_value=3),
    max_mob=st.integers(min_value=3, max_value=6),
    seed=st.integers(min_value=0, max_value=10000),
)
def test_property_8_portfolio_del_aggregation(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    seed: int
):
    """
    Property 8: Portfolio DEL Aggregation
    
    For any MOB, the Portfolio DEL SHALL equal the mean of all segment DELs at that MOB.
    
    **Validates: Requirements 6.1**
    """
    last_actual_mob = max_mob // 2
    
    mixed_wide, actual_wide, forecast_wide, flags_wide = create_del_dataframes(
        n_cohorts, n_segments, max_mob, last_actual_mob, seed
    )
    
    # Compute portfolio DEL
    portfolio_mixed, portfolio_actual, portfolio_forecast, portfolio_flags = \
        _compute_portfolio_del(mixed_wide, actual_wide, forecast_wide, flags_wide)
    
    mob_cols = [f"MOB_{i}" for i in range(max_mob + 1)]
    
    # Property check: Portfolio DEL = mean of segment DELs per cohort
    for cohort in mixed_wide["cohort"].unique():
        # Get segment DELs for this cohort
        segment_dels = mixed_wide[mixed_wide["cohort"] == cohort]
        
        # Get portfolio DEL for this cohort
        portfolio_row = portfolio_mixed[portfolio_mixed["cohort"] == cohort]
        
        assert len(portfolio_row) == 1, (
            f"Should have exactly one portfolio row for cohort {cohort}"
        )
        
        for mob_col in mob_cols:
            # Expected: mean of segment DELs
            expected_mean = segment_dels[mob_col].mean()
            
            # Actual: portfolio DEL
            actual_portfolio = portfolio_row[mob_col].values[0]
            
            assert np.isclose(actual_portfolio, expected_mean, rtol=1e-9, equal_nan=True), (
                f"Portfolio DEL at {mob_col} for cohort {cohort} should be {expected_mean}, "
                f"got {actual_portfolio}"
            )


# =============================================================================
# Property 9: Portfolio Flag Aggregation
# =============================================================================
# **Feature: del-forecast-calibration, Property 9: Portfolio Flag Aggregation**
# *For any* MOB, the Portfolio flag SHALL be "ACTUAL" if all segments have "ACTUAL",
# "MIXED" if some have "ACTUAL", and "FORECAST" if none have "ACTUAL".
# **Validates: Requirements 6.2**

@settings(max_examples=30, deadline=30000)
@given(
    n_cohorts=st.integers(min_value=1, max_value=3),
    n_segments=st.integers(min_value=2, max_value=3),
    max_mob=st.integers(min_value=3, max_value=6),
    seed=st.integers(min_value=0, max_value=10000),
)
def test_property_9_portfolio_flag_aggregation(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    seed: int
):
    """
    Property 9: Portfolio Flag Aggregation
    
    For any MOB, the Portfolio flag SHALL be "ACTUAL" if all segments have "ACTUAL",
    "MIXED" if some have "ACTUAL", and "FORECAST" if none have "ACTUAL".
    
    **Validates: Requirements 6.2**
    """
    last_actual_mob = max_mob // 2
    
    mixed_wide, actual_wide, forecast_wide, flags_wide = create_del_dataframes(
        n_cohorts, n_segments, max_mob, last_actual_mob, seed
    )
    
    # Compute portfolio flags
    _, _, _, portfolio_flags = \
        _compute_portfolio_del(mixed_wide, actual_wide, forecast_wide, flags_wide)
    
    mob_cols = [f"MOB_{i}" for i in range(max_mob + 1)]
    
    # Property check: Portfolio flag follows aggregation rules
    for cohort in flags_wide["cohort"].unique():
        # Get segment flags for this cohort
        segment_flags = flags_wide[flags_wide["cohort"] == cohort]
        
        # Get portfolio flags for this cohort
        portfolio_row = portfolio_flags[portfolio_flags["cohort"] == cohort]
        
        assert len(portfolio_row) == 1, (
            f"Should have exactly one portfolio row for cohort {cohort}"
        )
        
        for mob_col in mob_cols:
            # Get all segment flags for this MOB
            seg_flag_values = segment_flags[mob_col].values
            
            # Determine expected portfolio flag
            all_actual = all(f == "ACTUAL" for f in seg_flag_values)
            any_actual = any(f == "ACTUAL" for f in seg_flag_values)
            
            if all_actual:
                expected_flag = "ACTUAL"
            elif any_actual:
                expected_flag = "MIXED"
            else:
                expected_flag = "FORECAST"
            
            # Actual portfolio flag
            actual_flag = portfolio_row[mob_col].values[0]
            
            assert actual_flag == expected_flag, (
                f"Portfolio flag at {mob_col} for cohort {cohort} should be '{expected_flag}', "
                f"got '{actual_flag}'. Segment flags: {list(seg_flag_values)}"
            )


# =============================================================================
# Additional test: Mixed flags with varying last_actual_mob per segment
# =============================================================================

@settings(max_examples=30, deadline=30000)
@given(
    n_cohorts=st.integers(min_value=1, max_value=2),
    n_segments=st.integers(min_value=2, max_value=3),
    max_mob=st.integers(min_value=5, max_value=6),
    seed=st.integers(min_value=0, max_value=10000),
)
def test_property_9_mixed_flags_varying_actual_mob(
    n_cohorts: int,
    n_segments: int,
    max_mob: int,
    seed: int
):
    """
    Property 9: Portfolio Flag Aggregation with varying last_actual_mob per segment
    
    Tests the MIXED flag case where some segments have ACTUAL and others have FORECAST
    at the same MOB.
    
    **Validates: Requirements 6.2**
    """
    np.random.seed(seed)
    
    cohorts = [f"2024-{str(i+1).zfill(2)}" for i in range(n_cohorts)]
    segments = [f"SEG_{chr(65+i)}" for i in range(n_segments)]
    mob_cols = [f"MOB_{i}" for i in range(max_mob + 1)]
    
    # Create flags with varying last_actual_mob per segment
    records_mixed = []
    records_actual = []
    records_forecast = []
    records_flags = []
    
    for cohort in cohorts:
        for seg_idx, segment in enumerate(segments):
            # Different last_actual_mob for each segment
            last_actual_mob = 1 + seg_idx  # SEG_A: 1, SEG_B: 2, SEG_C: 3, etc.
            
            mixed_row = {"cohort": cohort, "segment_key": segment}
            actual_row = {"cohort": cohort, "segment_key": segment}
            forecast_row = {"cohort": cohort, "segment_key": segment}
            flags_row = {"cohort": cohort, "segment_key": segment}
            
            for mob in range(max_mob + 1):
                mob_col = f"MOB_{mob}"
                
                actual_val = np.random.uniform(0, 0.3) if mob <= last_actual_mob else np.nan
                forecast_val = np.random.uniform(0, 0.3)
                
                if mob <= last_actual_mob:
                    mixed_row[mob_col] = actual_val
                    flags_row[mob_col] = "ACTUAL"
                else:
                    mixed_row[mob_col] = forecast_val
                    flags_row[mob_col] = "FORECAST"
                
                actual_row[mob_col] = actual_val
                forecast_row[mob_col] = forecast_val
            
            records_mixed.append(mixed_row)
            records_actual.append(actual_row)
            records_forecast.append(forecast_row)
            records_flags.append(flags_row)
    
    mixed_wide = pd.DataFrame(records_mixed)
    actual_wide = pd.DataFrame(records_actual)
    forecast_wide = pd.DataFrame(records_forecast)
    flags_wide = pd.DataFrame(records_flags)
    
    # Compute portfolio flags
    _, _, _, portfolio_flags = \
        _compute_portfolio_del(mixed_wide, actual_wide, forecast_wide, flags_wide)
    
    # Property check
    for cohort in cohorts:
        segment_flags = flags_wide[flags_wide["cohort"] == cohort]
        portfolio_row = portfolio_flags[portfolio_flags["cohort"] == cohort]
        
        for mob in range(max_mob + 1):
            mob_col = f"MOB_{mob}"
            seg_flag_values = segment_flags[mob_col].values
            
            all_actual = all(f == "ACTUAL" for f in seg_flag_values)
            any_actual = any(f == "ACTUAL" for f in seg_flag_values)
            
            if all_actual:
                expected_flag = "ACTUAL"
            elif any_actual:
                expected_flag = "MIXED"
            else:
                expected_flag = "FORECAST"
            
            actual_flag = portfolio_row[mob_col].values[0]
            
            assert actual_flag == expected_flag, (
                f"Portfolio flag at {mob_col} for cohort {cohort} should be '{expected_flag}', "
                f"got '{actual_flag}'. Segment flags: {list(seg_flag_values)}"
            )


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
